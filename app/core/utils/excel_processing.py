"""Excel processing utilities for bulk upload operations."""

import io
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.logging import LoggerMixin


@dataclass
class HeaderStyle:
    """Configuration class for Excel header cell styling."""
    
    font: Optional[Font] = None
    fill: Optional[PatternFill] = None
    alignment: Optional[Alignment] = None
    
    @classmethod
    def create_default(cls) -> 'HeaderStyle':
        """Create default professional header style."""
        return cls(
            font=Font(bold=True, color="FFFFFF"),
            fill=PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            alignment=Alignment(horizontal="center", vertical="center")
        )
    
    @classmethod
    def create_custom(
        cls,
        bold: bool = True,
        font_color: str = "FFFFFF",
        background_color: str = "366092",
        horizontal_alignment: str = "center",
        vertical_alignment: str = "center"
    ) -> 'HeaderStyle':
        """
        Create custom header style with simplified parameters.
        
        Args:
            bold: Whether font should be bold
            font_color: Font color in hex format (without #)
            background_color: Background color in hex format (without #)
            horizontal_alignment: Horizontal alignment (left, center, right)
            vertical_alignment: Vertical alignment (top, center, bottom)
        
        Returns:
            HeaderStyle: Configured style object
        """
        return cls(
            font=Font(bold=bold, color=font_color),
            fill=PatternFill(start_color=background_color, end_color=background_color, fill_type="solid"),
            alignment=Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment)
        )


class ExcelProcessing(LoggerMixin):
    """Generic utility for processing Excel files for bulk upload operations."""

    def __init__(self):
        """Initialize the Excel processing utility."""
        pass

    #region PUBLIC METHODS - Main interface methods

    async def validate_excel_file(self, file: UploadFile) -> bool:
        """Validate that the uploaded file is an Excel file."""
        try:
            # Check file extension
            if not file.filename:
                self.log_error("File has no filename")
                return False
            
            file_extension = file.filename.lower().split('.')[-1]
            if file_extension not in ['xlsx', 'xls']:
                self.log_error("Invalid file extension", 
                             file_name=file.filename, 
                             extension=file_extension)
                return False
            
            # Check MIME type
            if file.content_type not in [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                'application/vnd.ms-excel',  # .xls
                'application/octet-stream'  # Sometimes Excel files have this MIME type
            ]:
                self.log_warning("Unexpected MIME type for Excel file", 
                               file_name=file.filename, 
                               content_type=file.content_type)
            
            # Additional validation: try to read a small portion of the file
            try:
                content = await file.read()
                if len(content) == 0:
                    self.log_error("Excel file is empty", file_name=file.filename)
                    return False
                
                # Try to read just the headers to validate file structure
                test_df = pd.read_excel(io.BytesIO(content), engine='openpyxl', nrows=0)
                self.log_info("Excel file structure validation passed", 
                             file_name=file.filename,
                             columns_count=len(test_df.columns))
                
                # Reset file pointer for later processing
                await file.seek(0)
                
            except Exception as validation_error:
                self.log_error("Excel file structure validation failed", 
                             file_name=file.filename,
                             error=str(validation_error))
                return False
            
            self.log_info("Excel file validation passed", file_name=file.filename)
            return True
            
        except Exception as e:
            self.log_error("Error validating Excel file", error=str(e))
            return False

    async def process_excel_file(
        self, 
        file: UploadFile,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, List[str]]] = None,
        field_types: Optional[Dict[str, str]] = None,
        required_fields: Optional[List[str]] = None
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Process Excel file and return parsed data and validation errors.
        
        Args:
            file: Uploaded Excel file to process
            sheet_name: Name of the Excel sheet to read (e.g., 'Plantilla SPF', 'Plantilla PMM')
                       If None, uses the default/first sheet
            column_mapping: Dictionary mapping field names to possible column names in Excel
                           If None, uses column names as-is from Excel
            field_types: Dictionary mapping field names to their data types (string, text, date, decimal)
                        If None, treats all fields as strings
            required_fields: List of field names that are required (cannot be empty/null)
                           If None, no required field validation is performed
        
        Returns:
            Tuple of (parsed_rows, validation_errors):
                - parsed_rows: List of dictionaries with processed row data
                - validation_errors: List of validation error messages
        """
        try:
            self.log_info("Starting Excel file processing", 
                         file_name=file.filename,
                         sheet_name=sheet_name,
                         has_column_mapping=column_mapping is not None,
                         has_field_types=field_types is not None,
                         has_required_fields=required_fields is not None)
            
            # Step 1: Read file content
            content = await file.read()
            
            # Step 2: Load DataFrame from Excel
            df = await self._load_excel_dataframe(content, file.filename, sheet_name)
            
            # Step 3: Log file information
            self._log_excel_info(df, file.filename, sheet_name)
            
            # Step 3.1: Validate Excel has data (not just empty)
            data_validation_errors = self._validate_excel_has_data(df)
            if data_validation_errors:
                # If no data, return early with validation errors
                self.log_warning("Excel file validation failed - no data found", 
                               validation_errors=data_validation_errors)
                return [], data_validation_errors
            
            # Step 3.2: Validate Excel has expected columns (if column mapping provided)
            column_validation_errors = []
            if column_mapping:
                column_validation_errors = self._validate_excel_columns(df, column_mapping)
                if column_validation_errors:
                    self.log_warning("Excel column validation failed", 
                                   validation_errors=column_validation_errors)
                    # Don't return early here, continue processing to show all errors
            
            # Step 4: Normalize column names (if column mapping is provided)
            if column_mapping:
                df_normalized = self._normalize_columns(df, column_mapping)
            else:
                # Use columns as-is from Excel
                df_normalized = df
                self.log_info("No column mapping provided, using Excel column names as-is",
                             columns=list(df.columns))
            
            # Step 5: Validate required fields (if specified)
            if required_fields:
                field_validation_errors = self._validate_required_fields(df_normalized, required_fields, column_mapping)
            else:
                field_validation_errors = []
            
            # Step 6: Process all rows (if field types are provided)
            if field_types:
                processed_rows, row_validation_errors = self._process_all_rows(df_normalized, field_types)
            else:
                # Convert DataFrame to list of dictionaries without type conversion
                processed_rows = df_normalized.to_dict('records')
                row_validation_errors = []
                self.log_info("No field types provided, returning data without type conversion")
            
            # Combine all validation errors
            all_validation_errors = column_validation_errors + field_validation_errors + row_validation_errors
            
            # Step 7: Log processing summary
            self._log_processing_summary(file.filename, df, processed_rows, all_validation_errors)
            
            return processed_rows, all_validation_errors
            
        except Exception as e:
            error_msg = f"Error processing Excel file: {str(e)}"
            self.log_error("Excel processing failed", error=str(e), file_name=file.filename)
            return [], [error_msg]

    async def get_excel_sheets(self, file: UploadFile) -> List[str]:
        """
        Get list of sheet names in the Excel file.
        
        Args:
            file: Uploaded Excel file
            
        Returns:
            List of sheet names
        """
        try:
            content = await file.read()
            
            # Try to get sheet names with openpyxl
            try:
                excel_file = pd.ExcelFile(io.BytesIO(content), engine='openpyxl')
                sheet_names = excel_file.sheet_names
                self.log_info("Excel sheets found", 
                             file_name=file.filename,
                             sheet_names=sheet_names)
                # Reset file pointer
                await file.seek(0)
                return sheet_names
            except Exception as openpyxl_error:
                self.log_warning("Error getting sheets with openpyxl, trying alternative", 
                               error=str(openpyxl_error))
                # Try with xlrd as fallback
                try:
                    excel_file = pd.ExcelFile(io.BytesIO(content), engine=None)
                    sheet_names = excel_file.sheet_names
                    self.log_info("Excel sheets found with alternative engine", 
                                 file_name=file.filename,
                                 sheet_names=sheet_names)
                    # Reset file pointer
                    await file.seek(0)
                    return sheet_names
                except Exception as fallback_error:
                    self.log_error("Failed to get sheet names with any engine",
                                 openpyxl_error=str(openpyxl_error),
                                 fallback_error=str(fallback_error),
                                 file_name=file.filename)
                    # Reset file pointer
                    await file.seek(0)
                    return []
                    
        except Exception as e:
            self.log_error("Error getting Excel sheet information", 
                          error=str(e), 
                          file_name=file.filename)
            return []

    async def create_excel_from_data(
        self, 
        headers: List[str], 
        rows_data: List[List[Any]], 
        sheet_name: str = "Data",
        apply_styling: bool = True,
        header_styles: Optional[Union[HeaderStyle, List[HeaderStyle], Dict[str, HeaderStyle]]] = None
    ) -> bytes:
        """
        Create Excel file from headers and rows data with customizable header styling.
        
        Args:
            headers: List of column headers
            rows_data: List of lists containing row data
            sheet_name: Name of the Excel sheet
            apply_styling: Whether to apply styling to headers
            header_styles: Custom styling for headers. Can be:
                         - HeaderStyle: Single style applied to all headers
                         - List[HeaderStyle]: List of styles (one per header, must match headers length)
                         - Dict[str, HeaderStyle]: Dictionary mapping header names to specific styles
                         - None: Uses default professional styling when apply_styling=True
            
        Returns:
            bytes: Excel file content as bytes
        """
        try:
            self.log_info("Creating Excel file from data", 
                         headers_count=len(headers),
                         rows_count=len(rows_data),
                         sheet_name=sheet_name,
                         apply_styling=apply_styling,
                         has_custom_styles=header_styles is not None)
            
            # Validate header_styles parameter if provided
            if header_styles is not None and apply_styling:
                header_styles = self._validate_and_normalize_header_styles(headers, header_styles)
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                
                if apply_styling:
                    # Apply styling to header
                    style = self._get_header_style_for_column(header, col - 1, header_styles)
                    self._apply_style_to_cell(cell, style)
            
            # Write data rows
            for row_idx, row_data in enumerate(rows_data, 2):
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)
            
            # Auto-adjust column widths
            self._auto_adjust_column_widths(ws)
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            self.log_info("Excel file created successfully", 
                         file_size_bytes=len(buffer.getvalue()),
                         final_rows=len(rows_data) + 1,  # +1 for header
                         final_columns=len(headers))
            
            return buffer.getvalue()
            
        except Exception as e:
            self.log_error("Error creating Excel file from data", 
                          error=str(e),
                          headers_count=len(headers) if headers else 0,
                          rows_count=len(rows_data) if rows_data else 0)
            raise Exception(f"Failed to create Excel file: {str(e)}")

    def get_file_info(self, file: UploadFile) -> Dict[str, Any]:
        """Get basic information about the uploaded file."""
        return {
            'filename': file.filename,
            'content_type': file.content_type,
            'size_bytes': file.size if hasattr(file, 'size') else None,
        }

    async def _load_excel_dataframe(self, content: bytes, filename: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Load Excel content into a pandas DataFrame.
        
        Args:
            content: Excel file content as bytes
            filename: Name of the file for logging
            sheet_name: Optional sheet name to read
            
        Returns:
            pd.DataFrame: Loaded Excel data
            
        Raises:
            ValueError: If unable to read Excel file with any engine
        """
        try:
            if sheet_name:
                self.log_info(f"Reading Excel file with specific sheet: {sheet_name}")
                return pd.read_excel(io.BytesIO(content), engine='openpyxl', sheet_name=sheet_name)
            else:
                self.log_info("Reading Excel file with default sheet")
                return pd.read_excel(io.BytesIO(content), engine='openpyxl')
                
        except Exception as excel_error:
            self.log_error("Error reading Excel file with openpyxl", 
                         error=str(excel_error), 
                         file_name=filename,
                         sheet_name=sheet_name)
            
            # Try with alternative engine as fallback
            return await self._load_excel_with_fallback(content, filename, sheet_name, excel_error)

    async def _load_excel_with_fallback(self, content: bytes, filename: str, sheet_name: Optional[str], primary_error: Exception) -> pd.DataFrame:
        """
        Attempt to load Excel with fallback engine.
        
        Args:
            content: Excel file content as bytes
            filename: Name of the file for logging
            sheet_name: Optional sheet name to read
            primary_error: Error from primary engine attempt
            
        Returns:
            pd.DataFrame: Loaded Excel data
            
        Raises:
            ValueError: If unable to read Excel file with any engine
        """
        try:
            if sheet_name:
                df = pd.read_excel(io.BytesIO(content), engine=None, sheet_name=sheet_name)
            else:
                df = pd.read_excel(io.BytesIO(content), engine=None)  # Let pandas choose
            
            self.log_info("Successfully read Excel with alternative engine")
            return df
            
        except Exception as fallback_error:
            self.log_error("Failed to read Excel with any engine", 
                         openpyxl_error=str(primary_error),
                         fallback_error=str(fallback_error),
                         file_name=filename,
                         sheet_name=sheet_name)
            raise ValueError(f"Unable to read Excel file. Primary error: {str(primary_error)}. Fallback error: {str(fallback_error)}")

    def _log_excel_info(self, df: pd.DataFrame, filename: str, sheet_name: Optional[str]) -> None:
        """
        Log information about the loaded Excel DataFrame.
        
        Args:
            df: Loaded DataFrame
            filename: Name of the file
            sheet_name: Name of the sheet (if specified)
        """
        self.log_info("Excel file loaded successfully", 
                     file_name=filename, 
                     sheet_name=sheet_name,
                     total_rows=len(df),
                     total_columns=len(df.columns))
        
        # Log column names for debugging
        self.log_debug("Excel columns found", 
                      columns=list(df.columns),
                      file_name=filename,
                      sheet_name=sheet_name)

    def _process_all_rows(self, df: pd.DataFrame, field_types: Dict[str, str]) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Process all rows in the DataFrame.
        
        Args:
            df: DataFrame with normalized columns
            field_types: Dictionary mapping field names to their types
            
        Returns:
            Tuple of (processed_rows, validation_errors)
        """
        processed_rows = []
        validation_errors = []
        
        for index, row in df.iterrows():
            try:
                processed_row = self._process_row(row, index + 1, field_types)  # +1 for Excel row number
                processed_rows.append(processed_row)
            except Exception as e:
                error_msg = f"Error al procesar la fila {index + 1}: {str(e)}"
                validation_errors.append(error_msg)
                self.log_warning("Row processing error", 
                               row_number=index + 1, 
                               error=str(e))
        
        return processed_rows, validation_errors

    def _log_processing_summary(self, filename: str, df: pd.DataFrame, processed_rows: List[Dict[str, Any]], validation_errors: List[str]) -> None:
        """
        Log summary of the processing results.
        
        Args:
            filename: Name of the processed file
            df: Original DataFrame
            processed_rows: Successfully processed rows
            validation_errors: List of validation errors
        """
        self.log_info("Excel processing completed", 
                     file_name=filename,
                     total_rows=len(df),
                     valid_rows=len(processed_rows),
                     invalid_rows=len(validation_errors))

    #endregion

    #region PRIVATE METHODS - Column Processing and Data Mapping

    def _normalize_columns(self, df: pd.DataFrame, column_mapping: Dict[str, List[str]]) -> pd.DataFrame:
        """Normalize column names to match expected field names."""
        try:
            mapping = {}
            found_columns = []
            missing_columns = []
            
            self.log_debug("Starting column normalization", 
                          excel_columns=list(df.columns),
                          expected_fields=list(column_mapping.keys()))
            
            for expected_field, possible_names in column_mapping.items():
                column_found = False
                for col_name in df.columns:
                    if str(col_name).strip() in possible_names:
                        mapping[col_name] = expected_field
                        found_columns.append(f"{col_name} -> {expected_field}")
                        column_found = True
                        break
                
                if not column_found:
                    missing_columns.append(f"{expected_field} (looking for: {possible_names})")
            
            # Rename columns
            df_renamed = df.rename(columns=mapping)
            
            self.log_info("Column normalization completed", 
                         mapped_columns=len(mapping),
                         expected_fields=len(column_mapping),
                         found_columns=found_columns,
                         missing_columns=missing_columns)
            
            if missing_columns:
                self.log_warning("Some expected columns were not found in Excel", 
                               missing_columns=missing_columns)
            
            return df_renamed
            
        except Exception as e:
            self.log_error("Error normalizing columns", error=str(e))
            raise

    #endregion

    #region PRIVATE METHODS - Data Validation

    def _validate_excel_has_data(self, df: pd.DataFrame) -> List[str]:
        """
        Validate that the Excel file has data rows (not just headers).
        
        Args:
            df: DataFrame to validate
            
        Returns:
            List of validation error messages
        """
        validation_errors = []
        
        try:
            if df.empty:
                validation_errors.append("El archivo Excel está vacío - no se han encontrado datos")
                self.log_warning("Excel file validation: file is completely empty")
                return validation_errors
            
            if len(df) == 0:
                validation_errors.append("El archivo Excel contiene solo encabezados pero no filas de datos")
                self.log_warning("Excel file validation: only headers found, no data rows")
                return validation_errors
            
            # Check if all data rows are empty (only NaN/None values)
            non_null_count = df.count().sum()  # Count non-null values across all columns
            if non_null_count == 0:
                validation_errors.append("El archivo Excel no contiene datos reales - todas las celdas de datos están vacías")
                self.log_warning("Excel file validation: all data cells are empty", 
                               total_rows=len(df),
                               total_columns=len(df.columns))
                return validation_errors
            
            self.log_info("Excel data validation passed", 
                         total_rows=len(df),
                         total_columns=len(df.columns),
                         non_null_values=non_null_count)
            
            return validation_errors
            
        except Exception as e:
            self.log_error("Error validating Excel data", error=str(e))
            return [f"Error during Excel data validation: {str(e)}"]

    def _validate_excel_columns(self, df: pd.DataFrame, column_mapping: Dict[str, List[str]]) -> List[str]:
        """
        Validate that the Excel file contains the expected column headers.
        
        Args:
            df: DataFrame to validate
            column_mapping: Dictionary mapping field names to possible column names in Excel
            
        Returns:
            List of validation error messages
        """
        validation_errors = []
        
        try:
            if df.empty:
                validation_errors.append("No se pueden validar columnas - El archivo Excel está vacío")
                return validation_errors
            
            excel_columns = [str(col).strip() for col in df.columns]
            
            missing_columns = []
            found_columns = []
            
            self.log_debug("Validating Excel columns", 
                          excel_columns=excel_columns,
                          expected_mappings=len(column_mapping))
            
            for expected_field, possible_names in column_mapping.items():
                column_found = False
                for possible_name in possible_names:
                    if possible_name in excel_columns:
                        found_columns.append(f"{possible_name} ({expected_field})")
                        column_found = True
                        break
                
                if not column_found:
                    # Only show the user-friendly column names in the error
                    missing_columns.append(', '.join(possible_names))
            
            if missing_columns:
                validation_errors.append(f"Faltan columnas necesarias en Excel: {'; '.join(missing_columns)}")
                self.log_warning("Excel column validation failed", 
                               missing_columns=missing_columns,
                               found_columns=found_columns)
            else:
                self.log_info("Excel column validation passed", 
                             found_columns=found_columns,
                             total_expected=len(column_mapping))
            
            return validation_errors
            
        except Exception as e:
            self.log_error("Error validating Excel columns", error=str(e))
            return [f"Error during Excel column validation: {str(e)}"]

    def _validate_required_fields(self, df: pd.DataFrame, required_fields: List[str], column_mapping: Optional[Dict[str, List[str]]] = None) -> List[str]:
        """
        Validate that required fields have values in all rows.
        
        Args:
            df: DataFrame with normalized columns
            required_fields: List of field names that are required
            column_mapping: Optional column mapping to get user-friendly field names for errors
            
        Returns:
            List of validation error messages
        """
        validation_errors = []
        
        try:
            # Create reverse mapping from technical field names to user-friendly names
            field_to_display_name = {}
            if column_mapping:
                for technical_field, possible_names in column_mapping.items():
                    # Use the first possible name as the display name (usually the main/primary name)
                    field_to_display_name[technical_field] = possible_names[0] if possible_names else technical_field
            
            self.log_info("Starting required fields validation", 
                         required_fields=required_fields,
                         total_rows=len(df),
                         has_display_names=bool(column_mapping))
            
            # Check each row for required fields
            for index, row in df.iterrows():
                row_number = index + 1  # +1 for Excel row numbering
                missing_fields = []
                
                for field in required_fields:
                    if field in row:
                        value = row[field]
                        # Check if value is None, NaN, or empty string
                        if pd.isna(value) or (isinstance(value, str) and value.strip() == ''):
                            # Use user-friendly name if available, otherwise use technical name
                            display_name = field_to_display_name.get(field, field)
                            missing_fields.append(display_name)
                    else:
                        # Field column not found in DataFrame
                        display_name = field_to_display_name.get(field, field)
                        missing_fields.append(display_name)
                
                if missing_fields:
                    error_msg = f"Fila {row_number}: Faltan campos requeridos: {', '.join(missing_fields)}"
                    validation_errors.append(error_msg)
            
            if validation_errors:
                self.log_warning("Required fields validation found issues", 
                               total_errors=len(validation_errors),
                               sample_errors=validation_errors[:5])  # Log first 5 errors as sample
            else:
                self.log_info("Required fields validation passed", 
                             total_rows=len(df),
                             required_fields_count=len(required_fields))
            
            return validation_errors
            
        except Exception as e:
            self.log_error("Error validating required fields", error=str(e))
            return [f"Error during required fields validation: {str(e)}"]

    #endregion

    #region PRIVATE METHODS - Row and Field Processing

    def _process_row(self, row: pd.Series, row_number: int, field_types: Dict[str, str]) -> Dict[str, Any]:
        """Process a single row and convert data types based on field type configuration."""
        processed_row = {}
        
        try:
            # Group fields by type for processing
            field_groups = self._group_fields_by_type(field_types)
            
            # Process each field type group
            for field_type, fields in field_groups.items():
                for field in fields:
                    processed_row[field] = self._process_field_value(row, field, field_type)
            
            return processed_row
            
        except Exception as e:
            self.log_error("Error processing row data", 
                          row_number=row_number, 
                          error=str(e))
            raise

    def _group_fields_by_type(self, field_types: Dict[str, str]) -> Dict[str, List[str]]:
        """
        Group fields by their data types.
        
        Args:
            field_types: Dictionary mapping field names to their types
            
        Returns:
            Dictionary with field types as keys and lists of field names as values
        """
        groups = {
            'string': [],
            'text': [],
            'date': [],
            'decimal': []
        }
        
        for field, field_type in field_types.items():
            if field_type in groups:
                groups[field_type].append(field)
        
        return groups

    def _process_field_value(self, row: pd.Series, field: str, field_type: str) -> Any:
        """
        Process a single field value based on its type.
        
        Args:
            row: Pandas Series representing the row
            field: Field name to process
            field_type: Type of the field (string, text, date, decimal)
            
        Returns:
            Processed field value
        """
        value = row.get(field)
        
        if pd.isna(value):
            return None
        
        if field_type in ['string', 'text']:
            return self._process_string_field(value)
        elif field_type == 'date':
            return self._process_date_field(value)
        elif field_type == 'decimal':
            return self._process_decimal_field(value)
        else:
            # Default: convert to string
            return str(value).strip() if value is not None else None

    #endregion

    #region PRIVATE METHODS - Field Type Processors

    def _process_string_field(self, value: Any) -> Optional[str]:
        """Process string/text field value."""
        if pd.notna(value):
            return str(value).strip()
        return None

    def _process_date_field(self, value: Any) -> Optional[datetime]:
        """Process date field value."""
        if pd.notna(value):
            try:
                if isinstance(value, str):
                    # Try to parse string date
                    return pd.to_datetime(value).to_pydatetime()
                else:
                    return pd.to_datetime(value).to_pydatetime()
            except:
                return None
        return None

    def _process_decimal_field(self, value: Any) -> Optional[Decimal]:
        """Process decimal field value."""
        if pd.notna(value):
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                return None
        return None

    #endregion

    #region PRIVATE METHODS - Excel Creation and Styling Utilities

    def _validate_and_normalize_header_styles(
        self, 
        headers: List[str], 
        header_styles: Union[HeaderStyle, List[HeaderStyle], Dict[str, HeaderStyle]]
    ) -> Union[HeaderStyle, List[HeaderStyle], Dict[str, HeaderStyle]]:
        """
        Validate and normalize header styles parameter.
        
        Args:
            headers: List of column headers
            header_styles: Header styles to validate
            
        Returns:
            Validated and normalized header styles
            
        Raises:
            ValueError: If header_styles format is invalid
        """
        if isinstance(header_styles, HeaderStyle):
            # Single style for all headers - valid
            return header_styles
        
        elif isinstance(header_styles, list):
            # List of styles - must match headers length
            if len(header_styles) != len(headers):
                raise ValueError(
                    f"Length of header_styles list ({len(header_styles)}) must match "
                    f"length of headers list ({len(headers)})"
                )
            # Validate each style is HeaderStyle instance
            for i, style in enumerate(header_styles):
                if not isinstance(style, HeaderStyle):
                    raise ValueError(f"header_styles[{i}] must be HeaderStyle instance, got {type(style)}")
            return header_styles
        
        elif isinstance(header_styles, dict):
            # Dictionary mapping - validate keys exist in headers
            invalid_keys = set(header_styles.keys()) - set(headers)
            if invalid_keys:
                raise ValueError(f"header_styles contains invalid header names: {invalid_keys}")
            # Validate each value is HeaderStyle instance
            for header_name, style in header_styles.items():
                if not isinstance(style, HeaderStyle):
                    raise ValueError(f"header_styles['{header_name}'] must be HeaderStyle instance, got {type(style)}")
            return header_styles
        
        else:
            raise ValueError(f"header_styles must be HeaderStyle, List[HeaderStyle], or Dict[str, HeaderStyle], got {type(header_styles)}")

    def _get_header_style_for_column(
        self, 
        header_name: str, 
        column_index: int, 
        header_styles: Optional[Union[HeaderStyle, List[HeaderStyle], Dict[str, HeaderStyle]]]
    ) -> HeaderStyle:
        """
        Get the appropriate style for a specific column header.
        
        Args:
            header_name: Name of the header column
            column_index: Zero-based index of the column
            header_styles: Header styles configuration
            
        Returns:
            HeaderStyle: Style to apply to the header cell
        """
        if header_styles is None:
            # No custom styles provided, use default
            return HeaderStyle.create_default()
        
        elif isinstance(header_styles, HeaderStyle):
            # Single style for all headers
            return header_styles
        
        elif isinstance(header_styles, list):
            # List of styles - use by index
            return header_styles[column_index]
        
        elif isinstance(header_styles, dict):
            # Dictionary mapping - use specific style if available, otherwise default
            return header_styles.get(header_name, HeaderStyle.create_default())
        
        else:
            # Fallback to default (should not reach here after validation)
            return HeaderStyle.create_default()

    def _apply_style_to_cell(self, cell, style: HeaderStyle) -> None:
        """
        Apply HeaderStyle to an Excel cell.
        
        Args:
            cell: Openpyxl cell object
            style: HeaderStyle to apply
        """
        if style.font is not None:
            cell.font = style.font
        
        if style.fill is not None:
            cell.fill = style.fill
        
        if style.alignment is not None:
            cell.alignment = style.alignment

    def _auto_adjust_column_widths(self, worksheet) -> None:
        """
        Auto-adjust column widths based on content length.
        
        Args:
            worksheet: Openpyxl worksheet object
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set width with minimum of 8 and maximum of 50 characters
            adjusted_width = min(max(max_length + 2, 8), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    #endregion
